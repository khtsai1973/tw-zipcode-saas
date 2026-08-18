"""檔案讀寫 ETL：csv / xlsx / xls / ods（不依賴 pandas）。"""

from __future__ import annotations

import csv
from pathlib import Path

SUPPORTED_EXT = {".csv", ".xlsx", ".xls", ".ods"}
ZIP_COL = "3+3郵遞區號"
NORM_COL = "正規化地址"
STATUS_COL = "查詢狀態"
MAX_ROWS = 1000

NAME_ALIASES = {"名稱", "姓名", "收件人", "name", "Name", "客戶名稱", "聯絡人"}
ADDRESS_ALIASES = {"地址", "住址", "收件地址", "address", "Address", "寄送地址", "通訊地址"}


def detect_columns(columns: list[str]) -> dict[str, str | None]:
    name_col = next((c for c in columns if c in NAME_ALIASES or "名" in str(c)), None)
    addr_col = next(
        (c for c in columns if c in ADDRESS_ALIASES or "址" in str(c) or "address" in str(c).lower()),
        None,
    )
    return {"name": name_col, "address": addr_col}


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_table(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXT:
        raise ValueError(f"不支援的格式：{ext}")

    if ext == ".csv":
        last_err: Exception | None = None
        for enc in ("utf-8-sig", "utf-8", "big5", "cp950"):
            try:
                with path.open("r", encoding=enc, newline="") as f:
                    reader = csv.DictReader(f)
                    if not reader.fieldnames:
                        raise ValueError("CSV 沒有欄位列")
                    columns = [str(c) for c in reader.fieldnames]
                    rows = [{c: _cell_str(row.get(c)) for c in columns} for row in reader]
                    return columns, rows
            except UnicodeDecodeError as exc:
                last_err = exc
                continue
        raise ValueError(f"無法辨識 CSV 編碼：{last_err}")

    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise ValueError("Excel 沒有欄位列")
        columns = [_cell_str(c) or f"欄位{i+1}" for i, c in enumerate(header)]
        rows: list[dict[str, str]] = []
        for values in rows_iter:
            rows.append({columns[i]: _cell_str(values[i] if i < len(values) else "") for i in range(len(columns))})
        return columns, rows

    if ext == ".xls":
        import xlrd

        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        if sheet.nrows < 1:
            raise ValueError("Excel 沒有欄位列")
        columns = [_cell_str(sheet.cell_value(0, c)) or f"欄位{c+1}" for c in range(sheet.ncols)]
        rows = []
        for r in range(1, sheet.nrows):
            rows.append({columns[c]: _cell_str(sheet.cell_value(r, c)) for c in range(sheet.ncols)})
        return columns, rows

    if ext == ".ods":
        from odf.opendocument import load
        from odf.table import Table, TableCell, TableRow
        from odf.text import P

        doc = load(path)
        tables = doc.getElementsByType(Table)
        if not tables:
            raise ValueError("ODS 沒有工作表")
        table = tables[0]

        def cell_text(cell) -> str:
            parts = []
            for p in cell.getElementsByType(P):
                parts.append("".join(node.data for node in p.childNodes if hasattr(node, "data")))
            return "".join(parts).strip()

        matrix: list[list[str]] = []
        for row in table.getElementsByType(TableRow):
            values = [cell_text(c) for c in row.getElementsByType(TableCell)]
            matrix.append(values)
        if not matrix:
            raise ValueError("ODS 沒有資料")
        width = max(len(r) for r in matrix)
        columns = [(matrix[0][i] if i < len(matrix[0]) and matrix[0][i] else f"欄位{i+1}") for i in range(width)]
        rows = []
        for r in matrix[1:]:
            rows.append({columns[i]: (r[i] if i < len(r) else "") for i in range(width)})
        return columns, rows

    raise ValueError(f"不支援的格式：{ext}")


def write_table(columns: list[str], rows: list[dict[str, str]], path: Path) -> None:
    ext = path.suffix.lower()
    path.parent.mkdir(parents=True, exist_ok=True)

    if ext == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            for row in rows:
                writer.writerow({c: row.get(c, "") for c in columns})
        return

    if ext in {".xlsx", ".xls"}:
        from openpyxl import Workbook

        target = path if ext == ".xlsx" else path.with_suffix(".xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(columns)
        for row in rows:
            ws.append([row.get(c, "") for c in columns])
        wb.save(target)
        return

    if ext == ".ods":
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableCell, TableRow
        from odf.text import P

        doc = OpenDocumentSpreadsheet()
        table = Table(name="Sheet1")

        def add_row(values: list[str]) -> None:
            tr = TableRow()
            for value in values:
                cell = TableCell()
                cell.addElement(P(text=value))
                tr.addElement(cell)
            table.addElement(tr)

        add_row(columns)
        for row in rows:
            add_row([row.get(c, "") for c in columns])
        doc.spreadsheet.addElement(table)
        doc.save(str(path))
        return

    raise ValueError(f"不支援的輸出格式：{ext}")


def resolve_output_path(original_name: str, job_id: str, output_dir: Path) -> Path:
    src = Path(original_name)
    ext = src.suffix.lower()
    if ext == ".xls":
        ext = ".xlsx"
    return output_dir / f"{job_id}_{src.stem}_with_zip{ext}"
